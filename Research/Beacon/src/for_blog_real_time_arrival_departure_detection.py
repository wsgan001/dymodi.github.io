# This is for the blog "Real-Time Arrival Departure Detection"
# 02/04/18 Yi DING

# Import
import openpyxl
import os
from datetime import datetime
from pytz import timezone
import matplotlib.pyplot as plt
from scipy import signal
import numpy as np


# # Draw RSSI - Distance relation
# rssi_list = []
# dist_list = []
# A = -55
# n = 2
# for rssi in range(-100, -50):
#     rssi_list.append(rssi)
#     dist = np.power(10, (-rssi+A)/(10*n))
#     dist_list.append(dist)
#
# # Plot In region result after shave
# plt.plot(dist_list, rssi_list)
# plt.xlabel('distance / meters')
# plt.ylabel('RSSI / dB ')
# plt.title('RSSI - Distance relation')
# plt.show()


# This path of current file
dir_path = os.path.dirname(os.path.realpath(__file__))
data_path = os.path.join(dir_path, '../data/')

# Read raw beacon data for specific (rider, shop) pair
rider_id = 100556090
shop_id = 147525716
date_str = '2018-02-04'

# Set file name
file_name = '_'.join(['dw_ai_clairvoyant_beacon', date_str, str(rider_id), str(shop_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)

# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row
column_count = sheet.max_column

# Get argument list
argument_list = []
for j in range(1, column_count+1):
    argument_list.append(sheet.cell(row=1,column=j).value)

print('argument list: ', argument_list)

# Plot whole timestamp - rssi figure
rssi_list = []
timestamp_list = []
rssi_index = argument_list.index('rssi')
timestamp_index = argument_list.index('unix_timestamp')
shop_index = argument_list.index('shop_id')
for i in range(2, row_count+1):
    rssi_list.append(sheet.cell(row=i, column=rssi_index+1).value)
    timestamp_list.append(int(sheet.cell(row=i, column=timestamp_index+1).value))

# Time conversion
date_str_start = "2018-02-04 11:00:00"
date_str_end = '2018-02-04 11:10:00'
time_start = datetime.strptime(date_str_start, "%Y-%m-%d %H:%M:%S")
time_end = datetime.strptime(date_str_end, "%Y-%m-%d %H:%M:%S")
time_start_SH = timezone('Asia/Shanghai').localize(time_start)
time_end_SH = timezone('Asia/Shanghai').localize(time_end)
timestamp_start =int(time_start_SH.timestamp())
timestamp_end =int(time_end_SH.timestamp())
# Pick RSSI from specific time slot
rssi_draw = []
timestamp_draw = []
for timestp in range(timestamp_start, timestamp_end):
    timestamp_draw.append(timestp)
    if timestp in timestamp_list:
        rssi_draw.append(rssi_list[timestamp_list.index(timestp)])
    else:
        rssi_draw.append(-100)

# # Plot Original RSSI in the time slot
# plt.plot(timestamp_draw, rssi_draw, 'C0')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop ')
# plt.show()

# Here we design a low-pass filter
fs = 1
nyq = 0.5 * fs
wn = 0.05 / nyq
order = 6
b, a = signal.butter(order, wn, 'low', analog = False)
rssi_filtered = signal.filtfilt(b, a, rssi_draw)

# # Plot filtered RSSI
# plt.plot(timestamp_draw, rssi_filtered, 'C1')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop (Offline filtered) ')
# plt.show()

# Plot Arrival/Departure Detection with RSSI threshold
in_region = []
rssi_thresh = -85
for rssi in rssi_filtered:
    if rssi > rssi_thresh:
        in_region.append(1)
    else:
        in_region.append(0)

# # Plot In region result
# plt.plot(timestamp_draw, in_region, 'C2')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information (Offline recognition)')
# plt.show()


# Binary signal x is filtered
# Zeros lasts less than n times are considered as noise
# Noise are removed
def window_filter(y, n):
    x = [i for i in y]
    headPos = 0
    tailPos = 0
    lastNbr = 0
    for currPos in range(0, len(x)):
        # Force set ones
        if lastNbr == 0 and x[currPos] == 1:
            headPos = currPos
            if currPos - tailPos < n and headPos <= tailPos:
                for i in range(tailPos, currPos):
                    x[i] = 1
        # Force set zeros
        if lastNbr == 1 and x[currPos] == 0:
            tailPos = currPos
            if currPos - headPos < n and headPos <= tailPos:
                for i in range(headPos, currPos):
                    x[i] = 0
        lastNbr = x[currPos]
    return x

# Shave
in_region = window_filter(in_region, 10)

# # Plot In region result after shave
# plt.plot(timestamp_draw, in_region, 'C3')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information after shave (Offline recognition)')
# plt.show()


# # Plot Butterworth filter frequency response
# w, h = signal.freqz(b, a)
# plt.semilogx(w, 20 * np.log10(abs(h)))
# plt.title('Butterworth filter frequency response')
# plt.xlabel('Frequency [radians / second]')
# plt.ylabel('Amplitude [dB]')
# plt.margins(0, 0.1)
# plt.grid(which='both', axis='both')
# plt.axvline(wn, color='green') # cutoff frequency
# plt.show()

# Design a moving filter and work on RSSI slot
horizon_length = 30
i = 0
rssi_range_filtered = []
while i + horizon_length < len(rssi_draw):
    rssi_window = rssi_draw[i:i+horizon_length]
    rssi_window_filtered = signal.filtfilt(b, a, rssi_window)
    for rssi in rssi_window_filtered:
        rssi_range_filtered.append(rssi)
    i = i + horizon_length

# # Plot RSSI with short range LPF
# len_filtered = len(rssi_range_filtered)
# plt.plot(timestamp_draw[0:len_filtered], rssi_range_filtered, 'C4')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\' RSSI value in 10 minutes for specific shop (Real-time filtered)')
# plt.show()

# Result of short range LPF after shaving
# Plot Arrival/Departure Detection with RSSI threshold
in_region = []
for rssi in rssi_range_filtered:
    if rssi > rssi_thresh:
        in_region.append(1)
    else:
        in_region.append(0)

# # Plot In region result
# plt.plot(timestamp_draw[0:len_filtered], in_region, 'C5')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information (Real-time recognition)')
# plt.show()

# Shave
in_region = window_filter(in_region, 10)

# # Plot In region result after shave
# plt.plot(timestamp_draw[0:len_filtered], in_region, 'C6')
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('In Region?')
# plt.title('Rider\'s in region information after shave (Real-time recognition--dede)')
# plt.show()


# Plot rider's RSSI with multiple shops
shop_id_list = [147525716, 153346655, 143531480, 866636, 155042113, 160260713, 2111770, 161311280, 159134974, 157030271,
                1522985, 161297653, 161397509, 1503512]

# Set file name
file_name = '_'.join(['dw_ai_clairvoyant_beacon', date_str, str(rider_id)])
file_name = '.'.join([file_name, 'xlsx'])
file_path = os.path.join(data_path, file_name)

# Open workbook
wb = openpyxl.load_workbook(file_path)
sheet = wb.get_sheet_by_name('Sheet0')
row_count = sheet.max_row

# Get three lists from raw file
rssi_list = []
timestamp_list = []
shop_list = []
for i in range(2, row_count+1):
    timestamp = int(sheet.cell(row=i, column=timestamp_index+1).value)
    if timestamp_start <= timestamp <= timestamp_end:
        shop_list.append(sheet.cell(row=i, column=shop_index+1).value)
        rssi_list.append(sheet.cell(row=i, column=rssi_index+1).value)
        timestamp_list.append(timestamp)

# Pick RSSI from specific time slot
time_axis = range(timestamp_start, timestamp_end+1)
rssi_matrix = np.full((len(time_axis), len(shop_id_list)), -100)
for i in range(0,len(rssi_list)):
    # What shop is it?
    shop_id = shop_list[i]
    shop_index = shop_id_list.index(shop_id)
    # What time is it?
    timestamp = timestamp_list[i]
    time_index = time_axis.index(timestamp)
    # Assign RSSI to the position
    rssi_matrix[time_index][shop_index] = rssi_list[i]

# # Plot rider's RSSI with respect to multiple shops
# for i in range(0, len(shop_id_list)):
#     plt.plot(time_axis, rssi_matrix[:,i], label = 'shop'+str(i))
# plt.xlabel('unix_timestamp / second')
# plt.ylabel('RSSI / dB (absent set as -100dB)')
# plt.title('Rider\'s in region information (Real-time recognition)')
# plt.legend()
# plt.show()


## -----------------------------------------------------------------------------
## Conduct Bayesian Estimation on Rider's Prescence

# Import P(R|A)
likelihood_rssi = []
for i in range(-109, -49, 1):
    likelihood_rssi.append(i)
likelihood_p = [0.000843471, 0.000909981, 0.001529736, 0.001680896, 0.002530413, 0.002865988, 0.003742714, 0.004138753,
 0.005626164, 0.006841488, 0.00897284,  0.009637942, 0.013350424, 0.012972525, 0.015575495, 0.014843882,
 0.021473747, 0.019995405, 0.025131811, 0.01923356,  0.030473795, 0.027242001, 0.032626309, 0.031646794,
 0.035513459, 0.032184923, 0.044214211, 0.041414734, 0.037463419, 0.037124822, 0.038920599, 0.032762353,
 0.036991801, 0.032623286, 0.033221878, 0.028327327, 0.038899436, 0.027725712, 0.031531913, 0.024357874,
 0.028723366, 0.02119259,  0.022879532, 0.017347087, 0.014710862, 0.010197233, 0.006384986, 0.004102474,
 0.002497158, 0.002043679, 0.001200208, 0.000649987, 0.000492781, 0.00021767,  7.25567E-05, 5.13943E-05,
 4.53479E-05, 1.20928E-05, 6.04639E-06, 9.06958E-06]


posterior_matrix = np.empty([len(time_axis),len(shop_id_list)])
for i in range(0, len(time_axis)):
    for j in range(0, len(shop_id_list)):
        posterior_matrix[i,j] = 